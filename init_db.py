import pymongo
from pymongo import MongoClient
from openpyxl import load_workbook


def connect_to_users():
    connect_to = MongoClient("localhost", 27017)
    db = connect_to["jungle_db"]
    return db["USERS"]


def connect_to_posts():
    connect_to = MongoClient("localhost", 27017)
    db = connect_to["jungle_db"]
    return db["POSTS"]


def connect_to_likes():
    connect_to = MongoClient("localhost", 27017)
    db = connect_to["jungle_db"]
    return db["LIKES"]


def initialize_db():
    # USERS 컬렉션 초기화
    users_collection = connect_to_users()
    users_collection.delete_many({})

    # POSTS 컬렉션 초기화
    posts_collection = connect_to_posts()
    posts_collection.delete_many({})

    # LIKES 컬렉션 초기화
    likes_collection = connect_to_likes()
    likes_collection.delete_many({})

    # USERS 컬렉션에 데이터 넣기
    wb = load_workbook(
        r"C:\Users\User\Desktop\hey\Jungle_DB.xlsx")

    def excel_to_DB(collection, wb):
        ws = wb.active
        for x in range(2, ws.max_row + 1):
            db_data = {
                "id": ws.cell(row=x, column=1).value if ws.cell(row=x, column=1).value else "",
                "name": ws.cell(row=x, column=2).value if ws.cell(row=x, column=2).value else "",
                "password": ws.cell(row=x, column=3).value if ws.cell(row=x, column=3).value else "",
                "posts": [],
                "isRegistered": ws.cell(row=x, column=5).value if ws.cell(row=x, column=5).value else False,
            }
            collection.insert_one(db_data)

    excel_to_DB(users_collection, wb)

    # POSTS 컬렉션 기본 필드 생성
    posts_collection.insert_one({
        "userId": None,
        "category": "",
        "title": "",
        "link": "",
        "content": "",
        "date": None,
        "like": 0,
        "liked_by": []
    })
    # posts_collection.delete_many({})  # 바로 삭제하여 실제 데이터는 남기지 않습니다.

    # LIKES 컬렉션 기본 필드 생성
    likes_collection.insert_one({
        "post": None,
        "user": None
    })
    # likes_collection.delete_many({})  # 바로 삭제하여 실제 데이터는 남기지 않습니다.


if __name__ == "__main__":
    initialize_db()
