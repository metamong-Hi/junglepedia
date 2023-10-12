import pymongo
from openpyxl import load_workbook


def Connect_DB():
    connect_to = pymongo.MongoClient("localhost", 27017)
    db = connect_to["jungle_db"]
    collection = db["jungle_users"]
    return collection


def initialize_db():
    # 데이터베이스 연결
    collection = Connect_DB()

    # 기존 데이터 삭제
    collection.delete_many({})

    wb = load_workbook(
        r"C:\Users\User\Desktop\hey\Jungle_DB.xlsx")

    def excel_to_DB(collection, wb):
        ws = wb.active
        for x in range(2, ws.max_row + 1):
            db_data = {
                "id": ws.cell(row=x, column=1).value if ws.cell(row=x, column=1).value else "",
                "name": ws.cell(row=x, column=2).value if ws.cell(row=x, column=2).value else "",
                "password": ws.cell(row=x, column=3).value if ws.cell(row=x, column=3).value else "",
                "posts": [],
                "isRegistered": ws.cell(row=x, column=5).value if ws.cell(row=x, column=5).value else False,
            }
            collection.insert_one(db_data)

    excel_to_DB(collection, wb)


if __name__ == "__main__":
    # 데이터베이스 초기화
    initialize_db()
